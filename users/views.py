from django.shortcuts import render
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from .models import Profile
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from .forms import UserRegisterForm


def register(request):
    if request.method == 'POST':
        form = UserRegisterForm(request.POST)
        if form.is_valid():
            user = form.save()
            user.refresh_from_db()
            user.game = form.cleaned_data.get('game')
            user.username = form.cleaned_data.get('username')
            user.ingamename = form.cleaned_data.get('ingame_name')
            user.firstname = form.cleaned_data.get('firstname')
            user.lastname = form.cleaned_data.get('lastname')
            user.discordtag = form.cleaned_data.get('discordtag')
            user.state = form.cleaned_data.get('state')
            user.email = form.cleaned_data.get('email')
            user.region = form.cleaned_data.get('region')
            user.shirt_size = form.cleaned_data.get('shirt_size')
            team = form.cleaned_data.get('team')
            user.team = team.name
            user.birthday = form.cleaned_data.get('birthday')
            user.save()
            messages.success(request, f'Registration successful!')
    else:
        form = UserRegisterForm()
    return render(request, 'users/register.html', {'form': form})


@login_required
def profile(request):
    return render(request, 'users/profile.html')


@login_required
def export_to_excel(request):
    players_set = Profile.objects.all()

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )

    response['Content-Disposition'] = 'attachment; filename={date}-players.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Sway Players'
    columns = [
        'Game',
        'Team',
        'Roster name',
        'In game name',
        'Firstname',
        'Lastname',
        'Discord#',
        'Email',
        'birthday',
        'Region',
        'State',
        'Shirt Size',
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = Font(bold=True)
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        if column_title == 'email':
            column_dimensions.width = 40
        elif column_title == 'team':
            column_dimensions.width = 30
        else:
            column_dimensions.width = 20

    for player in players_set:
        row_num += 1

        row = [
            player.game,
            player.team,
            player.username,
            player.ingame_name,
            player.firstname,
            player.lastname,
            player.discordtag,
            player.email,
            player.birthday,
            player.region,
            player.state,
            player.shirt_size
        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response
